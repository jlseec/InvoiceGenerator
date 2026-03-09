import os
import math
from openpyxl import load_workbook, Workbook
from copy import copy
from openpyxl.styles import Font, Border, Alignment, Side
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor

############################Logo Setting
def add_footer_logo(ws, footer_target_start_row,
                    logo_path,
                    width_inch=1.88,
                    height_inch=1.68,
                    column_letter="A",
                    offset_x_px=0,
                    offset_y_px=0):

    PIXELS_PER_INCH = 96
    EMU_PER_PIXEL = 9525  # Excel internal unit

    width_px = int(width_inch * PIXELS_PER_INCH)
    height_px = int(height_inch * PIXELS_PER_INCH)

    logo = Image(logo_path)
    logo.width = width_px
    logo.height = height_px

    # Convert column letter to index (A=0)
    col_index = ord(column_letter.upper()) - ord("A")
    row_index = footer_target_start_row  # zero-based internally

    marker = AnchorMarker(
        col=col_index,
        colOff=offset_x_px * EMU_PER_PIXEL,
        row=row_index,
        rowOff=offset_y_px * EMU_PER_PIXEL
    )

    size = XDRPositiveSize2D(
        cx=width_px * EMU_PER_PIXEL,
        cy=height_px * EMU_PER_PIXEL
    )

    logo.anchor = OneCellAnchor(_from=marker, ext=size)

    ws.add_image(logo)


############################Formatting
def style_items(ws, start_row, end_row, col_desc, col_qty, col_unit, col_amount):
    for row in range(start_row, end_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue

            # Font
            old_font = cell.font
            cell.font = Font(
                name="Segoe UI",
                size=old_font.size,
                bold=old_font.bold,
                italic=old_font.italic,
                underline=old_font.underline,
            )

            # Alignment
            if col == col_desc:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col in (col_qty, col_unit, col_amount):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")


def style_total_row(ws, total_row, col_qty, col_amount):
    ws.cell(row=total_row, column=col_qty).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=col_amount).alignment = Alignment(horizontal="center", vertical="center")


def style_footer(ws, footer_start_row, footer_height):
    for r in range(footer_start_row, footer_start_row + footer_height):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue

            old_font = cell.font
            cell.font = Font(
                name="Segoe UI",
                size=old_font.size,
                bold=old_font.bold,
                italic=old_font.italic,
                underline=old_font.underline,
            )


############################Generate Invoice
def generate_invoice(invoice):
    # ==============================
    # LAYOUT CONSTANTS
    # ==============================
    # LOGO SETTING
    invoice_type = invoice.get("invoice_type", 1)
    if invoice_type == 1:
        logo_file = "skmei_logo.jpg"
    else:
        logo_file = "solidfame_logo.jpg"

    #Counting
    ITEM_START_ROW = 15
    ROW_STEP = 3
    ITEMS_PER_PAGE = 10

    COL_DESC = 1        # A
    COL_QTY = 5         # E
    COL_UNIT_PRICE = 6  # F
    COL_AMOUNT = 7      # G

    FOOTER_SRC_START_ROW = 34
    FOOTER_SRC_END_ROW = 43
    FOOTER_START_COL = 1
    FOOTER_END_COL = 7

    # ==============================
    # PAGINATION
    # ==============================
    total_items = len(invoice["items"])
    total_page = math.ceil(total_items / ITEMS_PER_PAGE)

    # ==============================
    # OUTPUT WORKBOOK
    # ==============================
    output_wb = Workbook()
    output_wb.remove(output_wb.active)

    # ==============================
    # GRAND TOTALS (ALL PAGES)
    # ==============================
    grand_total_qty = 0
    grand_total_price = 0

    # ==============================
    # PAGE LOOP
    # ==============================
    all_qty_refs = []
    all_amount_refs = []
    for page_index in range(total_page):
        current_page = page_index + 1

        # ==============================
        # SELECT TEMPLATE BASED ON TYPE
        # ==============================
        if invoice_type == 1:
            template_file = "invoice_template_skmei.xlsx"
        elif invoice_type == 2:
            template_file = "invoice_template_solidfame.xlsx"
        else:
            raise ValueError("Invalid invoice_type. Use 1 or 2.")

        temp_wb = load_workbook(template_file)
        ws = temp_wb.active

        # ==============================
        # HEADER
        # ==============================
        ws.cell(row=9, column=2).value = invoice["customer"]
        ws.cell(row=10, column=2).value = invoice["attn"]
        ws.cell(row=11, column=2).value = invoice["email"]
        ws.cell(row=9, column=7).value = f"{current_page}/{total_page}"
        ws.cell(row=10, column=7).value = invoice["invoice_no"]
        ws.cell(row=11, column=7).value = invoice["date"]
            

        # ==============================
        # PAGE ITEMS SLICE
        # ==============================
        start_idx = page_index * ITEMS_PER_PAGE
        end_idx = start_idx + ITEMS_PER_PAGE
        page_items = invoice["items"][start_idx:end_idx]

        # ==============================
        # FORCE UNMERGE FOOTER AREA (IMPORTANT)
        # ==============================
        for merged in list(ws.merged_cells.ranges):
            if (
                merged.min_row >= FOOTER_SRC_START_ROW
                and merged.max_row <= FOOTER_SRC_END_ROW
            ):
                try:
                    ws.unmerge_cells(str(merged))
                except KeyError:
                    pass


        # ==============================
        # CUT FOOTER
        # ==============================
        footer_height = FOOTER_SRC_END_ROW - FOOTER_SRC_START_ROW + 1
        footer_cells = []
        footer_row_heights = {}
        footer_merges = []

        for r in range(FOOTER_SRC_START_ROW, FOOTER_SRC_END_ROW + 1):
            footer_row_heights[r] = ws.row_dimensions[r].height
            row_data = []
            for c in range(FOOTER_START_COL, FOOTER_END_COL + 1):
                cell = ws.cell(row=r, column=c)
                row_data.append({
                    "value": cell.value,
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "fill": copy(cell.fill),
                })
            footer_cells.append(row_data)

        # Capture footer merges AFTER unmerge (clean state)
        for merged in list(ws.merged_cells.ranges):
            if (
                merged.min_row >= FOOTER_SRC_START_ROW
                and merged.max_row <= FOOTER_SRC_END_ROW
            ):
                footer_merges.append(merged)

        ws.delete_rows(FOOTER_SRC_START_ROW, footer_height)
        # ==============================
        # UNMERGE ITEM AREA
        # ==============================
        if page_items:
            item_end_row = ITEM_START_ROW + (len(page_items) - 1) * ROW_STEP + 1
        else:
            item_end_row = ITEM_START_ROW

        for merged in list(ws.merged_cells.ranges):
            if (
                merged.min_row >= ITEM_START_ROW
                and merged.max_row <= item_end_row
            ):
                try:
                    ws.unmerge_cells(str(merged))
                except KeyError:
                    pass

        # ==============================
        # WRITE ITEMS
        # ==============================
        current_row = ITEM_START_ROW
        last_item_row = ITEM_START_ROW

        for item in page_items:
            qty = item["qty"]
            
            ws.cell(row=current_row, column=COL_DESC).value = f"Model no.: {item['item']}"
            ws.cell(row=current_row + 1, column=COL_DESC).value = item["description"]
            ws.cell(row=current_row, column=COL_QTY).value = qty
            ws.cell(row=current_row, column=COL_UNIT_PRICE).value = item["unit_price"]
            ws.cell(row=current_row, column=COL_AMOUNT).value = f"=E{current_row}*F{current_row}"

            sheet_name = "Invoice" if current_page == 1 else f"Invoice ({current_page})"

            all_qty_refs.append(f"'{sheet_name}'!E{current_row}")
            all_amount_refs.append(f"'{sheet_name}'!G{current_row}")

            last_item_row = current_row
            current_row += ROW_STEP

        # ==============================
        # TOTAL ROW (LAST PAGE ONLY)
        # ==============================
        if current_page == total_page:
            total_row = last_item_row + ROW_STEP

            # UNMERGE ANY MERGE OVERLAPPING TOTAL ROW
            for merged in list(ws.merged_cells.ranges):
                if merged.min_row <= total_row <= merged.max_row:
                    try:
                        ws.unmerge_cells(str(merged))
                    except KeyError:
                        pass

            # SUM only the quantity rows (every ROW_STEP)
            qty_rows = []
            amount_rows = []

            r = ITEM_START_ROW
            while r <= last_item_row:
                qty_rows.append(f"E{r}")
                amount_rows.append(f"G{r}")
                r += ROW_STEP

            ws.cell(row=total_row, column=COL_QTY).value = f"=SUM({','.join(all_qty_refs)})"
            ws.cell(row=total_row, column=COL_AMOUNT).value = f"=SUM({','.join(all_amount_refs)})"
           
            for col in (COL_QTY, COL_AMOUNT):
                ws.cell(row=total_row, column=col).alignment = Alignment(horizontal="center")

            top = Side(style="thin")
            bottom_double = Side(style="double")

            for col in range(1, 8):
                ws.cell(row=total_row, column=col).border = Border(top=top)

            ws.cell(row=total_row, column=COL_QTY).border = Border(top=top, bottom=bottom_double)
            ws.cell(row=total_row, column=COL_AMOUNT).border = Border(top=top, bottom=bottom_double)


            # ==============================
            # PASTE FOOTER
            # ==============================
            footer_target_start_row = total_row + 2
            ws.insert_rows(footer_target_start_row, footer_height)

            for r_offset, row_data in enumerate(footer_cells):
                tr = footer_target_start_row + r_offset
                ws.row_dimensions[tr].height = footer_row_heights[FOOTER_SRC_START_ROW + r_offset]

                for c_offset, cell_data in enumerate(row_data):
                    tc = FOOTER_START_COL + c_offset
                    cell = ws.cell(row=tr, column=tc)
                    cell.value = cell_data["value"]
                    cell.font = cell_data["font"]
                    cell.border = cell_data["border"]
                    cell.alignment = cell_data["alignment"]
                    cell.number_format = cell_data["number_format"]
                    cell.fill = cell_data["fill"]

            for merged in footer_merges:
                shift = footer_target_start_row - FOOTER_SRC_START_ROW
                ws.merge_cells(
                    start_row=merged.min_row + shift,
                    start_column=merged.min_col,
                    end_row=merged.max_row + shift,
                    end_column=merged.max_col
                )

            # ==============================
            # APPLY FOOTER MERGES (EXPLICIT)
            # ==============================
            thin_top = Border(top=Side(style="thin"))

            # --- 1. Footer row 2 col 5 → row 3 col 7 (LEFT aligned)
            # Relative: rows 2–3, cols E–G
            ws.merge_cells(
                start_row=footer_target_start_row + 1,
                start_column=5,
                end_row=footer_target_start_row + 2,
                end_column=7
            )
            cell_1 = ws.cell(row=footer_target_start_row + 1, column=5)
            cell_1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


            # Merge cell for 廣東時刻美錶業有限公司
            ws.merge_cells(
                start_row=footer_target_start_row + 9,
                start_column=1,
                end_row=footer_target_start_row + 9,
                end_column=3
            )
            cell_2 = ws.cell(row=footer_target_start_row + 9, column=1)
            cell_2.alignment = Alignment(horizontal="center", vertical="center")
            cell_2.border = thin_top

            # Merge cell for customer
            ws.merge_cells(
                start_row=footer_target_start_row + 9,
                start_column=6,
                end_row=footer_target_start_row + 9,
                end_column=7
            )

            cell_3 = ws.cell(row=footer_target_start_row + 9, column=6)

            # Set customer name into F43:G43
            cell_3.value = invoice["customer"]

            cell_3.alignment = Alignment(horizontal="center", vertical="center")
            cell_3.border = thin_top

        # ==============================
        # APPLY FORMATTING (ALL PAGES)
        # ==============================

        # Always style full item block
        end_row = last_item_row + (ROW_STEP - 1)

        # If this page has total row, include it
        if current_page == total_page:
            end_row = total_row

        # Style item section
        style_items(
            ws,
            ITEM_START_ROW,
            end_row,
            COL_DESC,
            COL_QTY,
            COL_UNIT_PRICE,
            COL_AMOUNT
        )

        # Style footer separately (only last page has footer)
        if current_page == total_page:
            style_footer(ws, footer_target_start_row, footer_height)


        # ==============================
        # COPY PAGE INTO OUTPUT WB
        # ==============================
        out_ws = output_wb.create_sheet(
            title="Invoice" if current_page == 1 else f"Invoice ({current_page})"
        )

        for row in ws.iter_rows():
            for cell in row:
                new_cell = out_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy(cell.alignment)

        for r, dim in ws.row_dimensions.items():
            out_ws.row_dimensions[r].height = dim.height

        for c, dim in ws.column_dimensions.items():
            out_ws.column_dimensions[c].width = dim.width

        for merged in ws.merged_cells.ranges:
            out_ws.merge_cells(str(merged))
  

        # Add logo to final output sheet (based on invoice_type)
        if current_page == total_page:
            add_footer_logo(
                out_ws,
                footer_target_start_row+1,
                logo_path=logo_file,
                width_inch=1.5,
                height_inch=1.35,
                column_letter="A",
                offset_x_px=40
            )
            
        out_ws.page_setup = copy(ws.page_setup)
        out_ws.page_margins = copy(ws.page_margins)
        out_ws.print_options = copy(ws.print_options)
        out_ws.print_title_rows = ws.print_title_rows
        out_ws.print_title_cols = ws.print_title_cols
        out_ws.print_area = ws.print_area
        out_ws.freeze_panes = ws.freeze_panes

    # ==============================
    # OUTPUT DIRECTORY (Inside Folder-Invoice Output)
    # ==============================
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "Invoice Output")

    os.makedirs(output_dir, exist_ok=True)

    base_name = f"Invoice_{invoice['invoice_no']}"
    output_file = os.path.join(output_dir, base_name + ".xlsx")

    i = 1
    while os.path.exists(output_file):
        output_file = os.path.join(output_dir, f"{base_name} ({i}).xlsx")
        i += 1

    output_wb.save(output_file)
    print(f"Invoice generated successfully: {output_file}")
    return output_file

        


